import openpyxl
import json
import os
from PIL import Image

def main():
    file_path = "src/MHAA_Voters_List.xlsx"
    print(f"Loading '{file_path}'...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    data = []
    
    # We will output images to public/images
    images_dir = "public/images"
    os.makedirs(images_dir, exist_ok=True)

    print(f"Sheet has {sheet.max_row} rows.")
    
    # Let's inspect the headers from row 1
    headers = [cell.value for cell in sheet[1]]
    print("Headers:", headers)
    
    # Usually openpyxl image objects are in sheet._images
    print(f"Number of images in sheet: {len(sheet._images)}")
    
    image_dict = {}
    for img in sheet._images:
        # img.anchor._from might contain row/col
        try:
            # openpyxl 3.1.2 uses ToMarker and FromMarker, but sometimes just anchor.
            # let's try to get row/col
            row = img.anchor._from.row + 1 # 1-indexed
            col = img.anchor._from.col + 1
            image_dict[row] = img
        except Exception as e:
            pass

    print(f"Mapped {len(image_dict)} images to rows.")

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {
            "serial": row[0],
            "enrollmentNo": row[1],
            "name": row[2],
            "phone": row[3],
            "address": row[4],
            "image": None
        }
        
        # Check if we have an image for this row
        if row_idx in image_dict:
            img = image_dict[row_idx]
            img_format = img.format if hasattr(img, 'format') else 'png'
            if img_format == 'jpeg': img_format = 'jpg'
            
            # Use enrollment number as image filename, or row_idx
            enrollment = str(row_data.get("enrollmentNo") or f"row_{row_idx}")
            enrollment_clean = "".join(c if c.isalnum() else "_" for c in enrollment)
            
            out_filename = f"{enrollment_clean}.{img_format}"
            out_path = os.path.join(images_dir, out_filename)
            
            # Save the image
            with open(out_path, "wb") as f:
                f.write(img._data())
                
            row_data["image"] = out_filename
            
        data.append(row_data)

    # Filter out empty rows if any
    data = [d for d in data if d["name"] or d["enrollmentNo"]]

    # Output to public/members.json
    out_json = "public/members.json"
    with open(out_json, "w") as f:
        json.dump(data, f, indent=2)
        
    print(f"Successfully wrote {len(data)} records to {out_json}.")

if __name__ == "__main__":
    main()
